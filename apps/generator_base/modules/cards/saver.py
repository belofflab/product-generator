import os 
import json 
from typing import Optional
from openpyxl import load_workbook

from apps.generator_base.services.user_card import UserCardService

from config.settings import MEDIA_DIR


class Saver:
    @classmethod
    def __init__(cls, user_id: int):
        cls.user_id = user_id

    def _add_obj(cls, save_path: str, sheet, card, row_start, index):
        card_data = cls.__load_card(f"{save_path}/{card['user_card_id']}")
        if card_data is None:
            return
        data = {**card, **card_data}
        row = row_start + index

        option_columns = {col.value: col.column for col in sheet[1]}

        sheet.cell(row=row, column=option_columns["Номер карточки"]).value = "0"
        sheet.cell(row=row, column=option_columns["Наименование"]).value = data["name"]
        sheet.cell(row=row, column=option_columns["Описание"]).value = data["description"]
        sheet.cell(row=row, column=option_columns["Бренд"]).value = data["brand"]
        sheet.cell(row=row, column=option_columns["Предмет"]).value = data["category_name"]

        for option in data['options']:
            col_name = option['name']
            if col_name in option_columns:
                col_index = option_columns[col_name]
                sheet.cell(row=row, column=col_index).value = option['value']

    @classmethod
    def make_report(cls, card_id: Optional[int] = None) -> None:
        wb = cls._load_sheet(MEDIA_DIR / "extras/raw_wb.xlsx")
        save_path = MEDIA_DIR / f"user_cards/{cls.user_id}"
        sheet = wb.active
        row_start = 2
        if card_id is None:
            data = UserCardService().list_cards(user_id=cls.user_id, per_page=10**3)
            for index, card in enumerate(data["items"]):
                cls._add_obj(cls, save_path, sheet, card, row_start, index)
        else:
            card = UserCardService().retrive_card(card_id)
            cls._add_obj(cls, save_path, sheet, card, row_start, 0)
            save_path =  f"{save_path}/{card_id}"
        wb.save(f"{save_path}/report.xlsx")

    @classmethod
    def _load_sheet(cls, template_path: str):
        return load_workbook(template_path)
    
    @classmethod
    def __load_card(cls, card_path: str):
        full_card_path = f"{card_path}/card.json"
        if not os.path.exists(full_card_path):
            return None
        return json.loads(open(full_card_path, "r", encoding="utf-8").read())